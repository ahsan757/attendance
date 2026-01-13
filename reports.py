from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import asyncio

# ------------------- MongoDB Setup -------------------
client = AsyncIOMotorClient("mongodb+srv://Ahsan12:Ahsan12@botss.rvm4jx6.mongodb.net/")
DB_NAME = "attendance_database"
LOCAL_TZ = ZoneInfo("Asia/Karachi")

def get_local_now():
    return datetime.now(LOCAL_TZ)

def get_collection_name(branch_name, date=None):
    if date is None:
        date = get_local_now()
    return f"{branch_name}_{date.strftime('%d_%m_%Y')}"

# ------------------- Report Generation Functions -------------------

async def generate_daily_report(date=None, branch_name=None, format='excel'):
    """
    Generate daily attendance report
    
    Args:
        date: datetime object (default: today)
        branch_name: specific branch or None for all branches
        format: 'excel' or 'csv'
    
    Returns:
        filename: path to generated report
    """
    if date is None:
        date = get_local_now()
    
    db = client[DB_NAME]
    
    # Get all branches if none specified
    if branch_name is None:
        branches = await db["branches"].find().to_list(length=100)
        branch_names = [b["branch_name"] for b in branches]
    else:
        branch_names = [branch_name]
    
    # Collect data from all branches
    all_data = []
    for branch in branch_names:
        coll_name = get_collection_name(branch, date)
        records = await db[coll_name].find().to_list(length=1000)
        
        for record in records:
            all_data.append({
                'Branch': record.get('branch', branch),
                'Employee Name': record.get('name', 'Unknown'),
                'Check In': record.get('check_in', 'N/A'),
                'Check Out': record.get('check_out', 'N/A'),
                'Total Hours': record.get('total_hours', 0),
                'Status': 'Present' if record.get('present') else 'Absent',
                'Device IP': record.get('device_ip', 'N/A')
            })
    
    if not all_data:
        print(f"No attendance records found for {date.strftime('%d-%m-%Y')}")
        return None
    
    # Create DataFrame
    df = pd.DataFrame(all_data)
    
    # Generate filename
    date_str = date.strftime('%d_%m_%Y')
    branch_suffix = f"_{branch_name}" if branch_name else "_All_Branches"
    
    if format == 'csv':
        filename = f"attendance_report{branch_suffix}_{date_str}.csv"
        df.to_csv(filename, index=False)
        print(f"✓ CSV report generated: {filename}")
        return filename
    
    else:  # Excel format with formatting
        filename = f"attendance_report{branch_suffix}_{date_str}.xlsx"
        
        # Create Excel with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {date_str}"
        
        # Header
        ws['A1'] = f"Daily Attendance Report - {date.strftime('%d %B %Y')}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # Column headers
        headers = ['Branch', 'Employee Name', 'Check In', 'Check Out', 'Total Hours', 'Status', 'Device IP']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_num, row_data in enumerate(all_data, 4):
            ws.cell(row=row_num, column=1, value=row_data['Branch'])
            ws.cell(row=row_num, column=2, value=row_data['Employee Name'])
            ws.cell(row=row_num, column=3, value=row_data['Check In'])
            ws.cell(row=row_num, column=4, value=row_data['Check Out'])
            ws.cell(row=row_num, column=5, value=row_data['Total Hours'])
            
            # Status cell with color coding
            status_cell = ws.cell(row=row_num, column=6, value=row_data['Status'])
            if row_data['Status'] == 'Present':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            ws.cell(row=row_num, column=7, value=row_data['Device IP'])
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        
        # Summary section
        summary_row = len(all_data) + 5
        ws.cell(row=summary_row, column=1, value='Summary:').font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value='Total Employees:')
        ws.cell(row=summary_row + 1, column=2, value=len(all_data))
        ws.cell(row=summary_row + 2, column=1, value='Total Hours:')
        ws.cell(row=summary_row + 2, column=2, value=f'=SUM(E4:E{len(all_data)+3})')
        ws.cell(row=summary_row + 3, column=1, value='Average Hours:')
        ws.cell(row=summary_row + 3, column=2, value=f'=AVERAGE(E4:E{len(all_data)+3})')
        
        wb.save(filename)
        print(f"✓ Excel report generated: {filename}")
        return filename


async def generate_weekly_report(start_date=None, branch_name=None):
    """Generate weekly attendance summary"""
    if start_date is None:
        start_date = get_local_now() - timedelta(days=6)  # Last 7 days
    
    db = client[DB_NAME]
    
    # Get branches
    if branch_name is None:
        branches = await db["branches"].find().to_list(length=100)
        branch_names = [b["branch_name"] for b in branches]
    else:
        branch_names = [branch_name]
    
    # Collect data for 7 days
    weekly_data = []
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        
        for branch in branch_names:
            coll_name = get_collection_name(branch, current_date)
            records = await db[coll_name].find().to_list(length=1000)
            
            for record in records:
                weekly_data.append({
                    'Date': current_date.strftime('%d-%m-%Y'),
                    'Day': current_date.strftime('%A'),
                    'Branch': record.get('branch', branch),
                    'Employee': record.get('name', 'Unknown'),
                    'Check In': record.get('check_in', 'N/A'),
                    'Check Out': record.get('check_out', 'N/A'),
                    'Hours': record.get('total_hours', 0)
                })
    
    if not weekly_data:
        print("No records found for the week")
        return None
    
    df = pd.DataFrame(weekly_data)
    
    # Generate filename
    end_date = start_date + timedelta(days=6)
    filename = f"weekly_report_{start_date.strftime('%d_%m')}_{end_date.strftime('%d_%m_%Y')}.xlsx"
    
    # Create Excel with pivot summary
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Detailed data
        df.to_excel(writer, sheet_name='Detailed', index=False)
        
        # Employee summary
        summary = df.groupby('Employee').agg({
            'Hours': 'sum',
            'Date': 'count'
        }).rename(columns={'Date': 'Days Present'})
        summary.to_excel(writer, sheet_name='Employee Summary')
        
        # Branch summary
        branch_summary = df.groupby('Branch').agg({
            'Hours': 'sum',
            'Employee': 'nunique'
        }).rename(columns={'Employee': 'Total Employees'})
        branch_summary.to_excel(writer, sheet_name='Branch Summary')
    
    print(f"✓ Weekly report generated: {filename}")
    return filename


async def generate_monthly_report(year=None, month=None, branch_name=None):
    """Generate monthly attendance summary"""
    now = get_local_now()
    if year is None:
        year = now.year
    if month is None:
        month = now.month
    
    db = client[DB_NAME]
    
    # Get branches
    if branch_name is None:
        branches = await db["branches"].find().to_list(length=100)
        branch_names = [b["branch_name"] for b in branches]
    else:
        branch_names = [branch_name]
    
    # Get all days in month
    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]
    
    monthly_data = []
    for day in range(1, days_in_month + 1):
        current_date = datetime(year, month, day, tzinfo=LOCAL_TZ)
        
        for branch in branch_names:
            coll_name = get_collection_name(branch, current_date)
            records = await db[coll_name].find().to_list(length=1000)
            
            for record in records:
                monthly_data.append({
                    'Date': current_date.strftime('%d-%m-%Y'),
                    'Branch': record.get('branch', branch),
                    'Employee': record.get('name', 'Unknown'),
                    'Hours': record.get('total_hours', 0),
                    'Status': 'Present' if record.get('present') else 'Absent'
                })
    
    if not monthly_data:
        print(f"No records found for {month}/{year}")
        return None
    
    df = pd.DataFrame(monthly_data)
    
    # Generate filename
    month_name = datetime(year, month, 1).strftime('%B')
    filename = f"monthly_report_{month_name}_{year}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Detailed data
        df.to_excel(writer, sheet_name='Detailed', index=False)
        
        # Employee monthly summary
        employee_summary = df.groupby('Employee').agg({
            'Hours': 'sum',
            'Date': 'nunique',
            'Status': lambda x: (x == 'Present').sum()
        }).rename(columns={'Date': 'Days Worked', 'Status': 'Present Days'})
        employee_summary['Attendance %'] = (employee_summary['Present Days'] / days_in_month * 100).round(2)
        employee_summary.to_excel(writer, sheet_name='Employee Summary')
        
        # Branch summary
        branch_summary = df.groupby('Branch').agg({
            'Hours': 'sum',
            'Employee': 'nunique'
        }).rename(columns={'Employee': 'Total Employees'})
        branch_summary.to_excel(writer, sheet_name='Branch Summary')
    
    print(f"✓ Monthly report generated: {filename}")
    return filename


# ------------------- CLI Interface -------------------
async def main():
    import sys
    
    if len(sys.argv) < 2:
        print("""
Usage:
    python reports.py daily [branch_name] [format]
    python reports.py weekly [branch_name]
    python reports.py monthly [branch_name]

Examples:
    python reports.py daily                          # All branches, Excel format
    python reports.py daily Karachi_Clifton csv      # Specific branch, CSV
    python reports.py weekly Karachi_Clifton         # Weekly report
    python reports.py monthly                        # Monthly report, all branches
        """)
        return
    
    report_type = sys.argv[1]
    branch = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] not in ['csv', 'excel'] else None
    format_type = sys.argv[3] if len(sys.argv) > 3 else 'excel'
    
    if report_type == 'daily':
        await generate_daily_report(branch_name=branch, format=format_type)
    elif report_type == 'weekly':
        await generate_weekly_report(branch_name=branch)
    elif report_type == 'monthly':
        await generate_monthly_report(branch_name=branch)
    else:
        print(f"Unknown report type: {report_type}")

if __name__ == "__main__":
    asyncio.run(main())